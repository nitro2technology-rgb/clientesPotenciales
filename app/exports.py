"""Generacion de archivos Excel (.xlsx) y CSV para descarga."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Negocio
from app.sheets import ENCABEZADOS

ANCHOS = [18, 18, 18, 32, 40, 18, 34, 8, 32, 8, 10, 40, 22, 18, 30]

VERDE = PatternFill("solid", fgColor="D9EAD3")   # lead valido
GRIS = PatternFill("solid", fgColor="EFEFEF")    # ya tiene web
AZUL = PatternFill("solid", fgColor="1F3864")    # encabezado


def _filas(negocios: list[Negocio]) -> list[list]:
    return [
        [
            n.fecha_busqueda, n.ciudad_buscada, n.categoria_buscada, n.nombre,
            n.direccion, n.telefono, n.sitio_web, "Si" if n.es_lead else "No",
            n.motivo,
            n.rating if n.rating is not None else "",
            n.resenas if n.resenas is not None else "",
            n.maps_url, n.email, n.estado_contacto, n.place_id,
        ]
        for n in negocios
    ]


def a_csv(negocios: list[Negocio]) -> bytes:
    buffer = io.StringIO(newline="")
    escritor = csv.writer(buffer, delimiter=";")  # ; para Excel en espanol
    escritor.writerow(ENCABEZADOS)
    escritor.writerows(_filas(negocios))
    # BOM para que Excel respete las tildes
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def a_xlsx(negocios: list[Negocio], titulo: str = "Leads") -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo[:31] or "Leads"

    hoja.append(ENCABEZADOS)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = AZUL
        celda.alignment = Alignment(vertical="center")

    for negocio, fila in zip(negocios, _filas(negocios)):
        hoja.append(fila)
        relleno = VERDE if negocio.es_lead else GRIS
        for celda in hoja[hoja.max_row]:
            celda.fill = relleno
        # Link clicable a Google Maps
        celda_maps = hoja.cell(row=hoja.max_row, column=12)
        if negocio.maps_url:
            celda_maps.hyperlink = negocio.maps_url
            celda_maps.value = "Ver en Maps"
            celda_maps.font = Font(color="0563C1", underline="single")

    for indice, ancho in enumerate(ANCHOS, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(ENCABEZADOS))}{hoja.max_row}"

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
